from openpyxl.styles import Font, PatternFill, Alignment, Color
from datetime import datetime

# -----------------------------
# Conditional rules evaluator
# -----------------------------
def evaluate_condition(value, condition):
    try:
        if value is None:
            return False
        num = float(value)
        expr = condition.replace("value", str(num))
        return eval(expr, {"__builtins__": {}})
    except Exception:
        return False

def apply_conditional_rules(cell, value, rules):
    if not rules:
        return
    for rule in rules:
        condition = rule.get("condition")
        style = rule.get("style")
        if not condition or not style:
            continue
        if evaluate_condition(value, condition):
            font_cfg = style.get("font")
            if font_cfg:
                color = font_cfg.get("color", "000000")
                if len(color) == 6:
                    color = "FF" + color
                cell.font = Font(
                    bold=font_cfg.get("bold", cell.font.bold),
                    color=Color(rgb=color)
                )
            fill = style.get("fill")
            if fill:
                if len(fill) == 6:
                    fill = "FF" + fill
                cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            align = style.get("alignment")
            if align:
                cell.alignment = Alignment(horizontal=align, vertical="center")

# -----------------------------
# Detect cell type
# -----------------------------
def detect_cell_type(value, column_type=None):
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return "text"

    # Column hints have priority
    if column_type in ["text", "category"]:
        return "text"
    if column_type == "money":
        try:
            num = float(value)
            return "money_negative" if num < 0 else "money_positive"
        except:
            return "text"

    # Auto-detect
    if isinstance(value, (int, float)):
        return "money_negative" if value < 0 else "money_positive"
    if isinstance(value, str):
        value_strip = value.strip()
        # detect date
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                datetime.strptime(value_strip, fmt)
                return "date"
            except:
                continue
        # detect numeric
        try:
            num = float(value_strip)
            return "money_negative" if num < 0 else "money_positive"
        except:
            return "text"
    return "text"

# -----------------------------
# Detect column types based on header
# -----------------------------
def detect_column_types(sheet, column_hints):
    column_types = {}
    header_row = list(sheet.iter_rows(min_row=1, max_row=1))[0]
    for col_index, cell in enumerate(header_row, start=1):
        value = cell.value
        if not value:
            column_types[col_index] = "text"
            continue
        header = str(value).strip().lower()
        detected_type = "text"
        for keyword, col_type in column_hints.items():
            if keyword in header:
                detected_type = col_type
                break
        column_types[col_index] = detected_type
    return column_types

# -----------------------------
# Apply base style
# -----------------------------
def apply_cell_style(cell, style):
    if not style:
        return
    font_cfg = style.get("font")
    if font_cfg:
        color = font_cfg.get("color", "000000")
        if len(color) == 6:
            color = "FF" + color
        cell.font = Font(
            bold=font_cfg.get("bold", False),
            color=Color(rgb=color)
        )
    fill = style.get("fill")
    if fill:
        if len(fill) == 6:
            fill = "FF" + fill
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    align = style.get("alignment")
    if align:
        cell.alignment = Alignment(horizontal=align, vertical="center")

# -----------------------------
# Apply styles engine
# -----------------------------
def apply_styles(workbook, style_config, column_hints):
    for sheet in workbook.worksheets:
        # Freeze header
        sheet.freeze_panes = "A2"

        # Detect column types
        column_types = detect_column_types(sheet, column_hints)

        # Apply styles
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            for cell in row:
                col_index = cell.column
                column_type = column_types.get(col_index)

                style_key = "header" if row_index == 1 else detect_cell_type(cell.value, column_type)
                style = style_config.get(style_key)
                apply_cell_style(cell, style)

                rules = style_config.get("rules", [])
                apply_conditional_rules(cell, cell.value, rules)

        # Auto-width columns
        for col in sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
                except:
                    continue
            sheet.column_dimensions[column_letter].width = max_length + 2