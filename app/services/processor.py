import pandas as pd
from openpyxl import Workbook
from app.styler_engine import apply_styles
from app.style_config import STYLE_CONFIG, COLUMN_HINTS
from pathlib import Path
from app.infrastructure.db import increment_attempts, update_task, get_task
from app.core.constants import MAX_RETRIES
import logging

# ------------------------
# Настройки категорий
# ------------------------
REQIRED_COLUMNS = ["Дата", "Описание", "Сумма", "Контрагент"]

CATEGORIES = {
    "аренда": ["аренда", "офис"],
    "зарплата": ["зарплата", "сотрудник"],
    "транспорт": ["азс", "топливо", "бензин", "солярка", "заправка"],
    "комиссии": ["комиссия"],
    "связь": ["интернет", "связь"],
}

CATEGORY_PRIORITY = {
    "зарплата": 100,
    "аренда": 90,
    "комиссии": 50,
    "транспорт": 40,
}

# ------------------------
# Работа с CSV
# ------------------------
def load_data(file_path: str) -> pd.DataFrame:
    df = pd.read_csv(file_path)
    df.columns = df.columns.str.strip()
    missing = [col for col in REQIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Отсутствуют колонки: {missing}")
    return df

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(subset=["Сумма", "Описание"])
    df["Сумма"] = pd.to_numeric(df["Сумма"], errors="coerce")
    df = df.dropna(subset=["Сумма"])
    return df

def categorize(row):
    text = f"{row.get('Описание','')} {row.get('Контрагент','')}".lower()
    best_category = "прочее"
    best_score = 0
    for category, keywords in CATEGORIES.items():
        for word in keywords:
            if word in text:
                score = CATEGORY_PRIORITY.get(category, 10)
                if score > best_score:
                    best_score = score
                    best_category = category
    return best_category

def process_file(file_path: str) -> pd.DataFrame:
    df = load_data(file_path)
    df = clean_data(df)
    df["Категория"] = df.apply(categorize, axis=1)
    return df

# ------------------------
# Создаём сводку
# ------------------------
def build_summary(df: pd.DataFrame):
    income = float(df[df["Сумма"] > 0]["Сумма"].sum())
    expense = float(df[df["Сумма"] < 0]["Сумма"].sum())
    profit = income + expense

    by_category = df.groupby("Категория")["Сумма"].sum().to_dict()
    expense_by_category = {k: float(v) for k, v in by_category.items()}

    return {
        "доход": income,
        "расход": expense,
        "прибыль": profit,
        "расход по категориям": expense_by_category
    }

# ------------------------
# Экспорт в Excel
# ------------------------
def export_to_excel(df: pd.DataFrame, summary: dict, output_path: str):
    """
    Создаёт Excel файл с тремя листами и применяет стили:
    - Лист 'Операции' — все операции из df
    - Лист 'Сводка' — доход/расход/прибыль
    - Лист 'Категории' — расходы по категориям
    """
    wb = Workbook()

    # ------------------------
    # 1️⃣ Лист "Операции"
    # ------------------------
    ws_ops = wb.active
    ws_ops.title = "Операции"

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_ops.cell(row=r_idx, column=c_idx, value=value)

    # Добавим заголовки
    for c_idx, col_name in enumerate(df.columns, start=1):
        ws_ops.cell(row=1, column=c_idx, value=col_name)

    # ------------------------
    # 2️⃣ Лист "Сводка"
    # ------------------------
    ws_sum = wb.create_sheet("Сводка")
    ws_sum.append(["Показатель", "Сумма"])
    ws_sum.append(["Доход", summary.get("доход", 0)])
    ws_sum.append(["Расход", summary.get("расход", 0)])
    ws_sum.append(["Прибыль", summary.get("прибыль", 0)])

    # ------------------------
    # 3️⃣ Лист "Категории"
    # ------------------------
    ws_cat = wb.create_sheet("Категории")
    ws_cat.append(["Категория", "Сумма"])
    expense_by_category = summary.get("расход по категориям", {})
    for cat, amount in expense_by_category.items():
        ws_cat.append([cat, amount])

    # ------------------------
    # Применяем стили ко всем листам
    # ------------------------
    apply_styles(wb, STYLE_CONFIG, COLUMN_HINTS)

    # Сохраняем Excel
    wb.save(output_path)

def process_in_background(
    input_path: Path,
    output_path: Path,
    file_id: str,
    ):

    """Обрабатывает загруженный CSV файл в background"""

    attempts = increment_attempts(file_id)

    try:
        # Обрабатываем данные
        df = process_file(input_path)
        summary = build_summary(df)
        export_to_excel(df, summary, output_path)

        update_task(file_id, status="done", output_path=str(output_path))
        logging.info(f"[PROCESS] done file_id={file_id}")

    except Exception as e:
        if attempts < MAX_RETRIES:
            update_task(file_id, status="retry", error=str(e))
            logging.warning(f"[PROCESS] retry now file_id={file_id} attempts={attempts}")
            return "retry"

        else:
            update_task(file_id, status="failed", error=str(e))
            logging.error(f"[PROCESS] failed file_id={file_id} attempts={attempts}")


# ------------------------
# Пример использования
# ------------------------
if __name__ == "__main__":
    input_csv = "./files/001.csv"
    output_xlsx = "./files/report_styled.xlsx"

    # Загружаем и обрабатываем CSV
    df_operations = process_file(input_csv)

    # Создаём сводку и категории
    summary_df, categories_df = build_summary(df_operations)

    # Экспортируем всё в Excel с тремя листами
    export_to_excel(df_operations, summary_df, categories_df, output_xlsx)